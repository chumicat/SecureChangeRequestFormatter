'''
A script to convert all xlsx files in the same directory to TOS Aurora Secure Track Accept Format
'''
import glob
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from itertools import chain
from typing import Sequence
from termcolor import colored
import openpyxl


CONFIG_PATH = "config.json"
LOG_PATH = "output.log"
ENCODING = "utf8"
EXTS = ("*.xlsx", "*.xls")
ALL_CONFIGS = {"SRC", "DST", "SRV", "ADD", "RM", "USG", "CMT"}
REQUIRED_CONFIGS = {"SRC", "DST", "SRV", "ADD", "RM"}
COMPLEX_CONFIGS = {"service_replace"}


class TeeOutput:
    """Handle Standard Output and make write to both standard output and a log file"""
    def __init__(self, filename):
        self.terminal = sys.stdout
        self.log_file = open(filename, 'a', encoding=ENCODING)

    def write(self, message):
        """Write to standard output and log file"""
        # Write colored output to terminal
        self.terminal.write(message)
        # Write color code striped output to file
        cleaned_message = self.strip_ansi(message)
        if cleaned_message.strip():  # Only add timestamp for non-empty lines
            timestamp = datetime.now().strftime('[%Y-%m-%d %H:%M:%S.%f] ')
            self.log_file.write(timestamp + cleaned_message)
        else:
            self.log_file.write(cleaned_message)

    def flush(self):
        """Update standard output and log file"""
        self.terminal.flush()
        self.log_file.flush()

    def strip_ansi(self, text):
        """Remove ANSI color codes from text"""
        ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
        return ansi_escape.sub('', text)

    def __getattr__(self, attr):
        """Forward all other attributes to terminal"""
        return getattr(self.terminal, attr)


def start_logging(log_file="output.log"):
    """
    Start logging all stdout to both terminal and file with timestamps.
    
    Args:
        log_file (str): Path to the log file
        
    Return:
        None
    """
    sys.stdout = TeeOutput(LOG_PATH)


def get_configs(config_path: str) -> dict:
    '''
    Read config and Return a variable dict.
    
        Parameters:
            config_path (str): A filepath to the json format config file.

        Returns:
            configs (dict): A dict with configs read from the confg json file.
    '''
    try:
        with open(config_path, 'r', encoding=ENCODING) as config_file:
            return json.load(config_file)
    except FileNotFoundError:
        print(colored(f"ERROR: Configuration file '{config_path}' not found.", "red"))
        return {}
    except json.JSONDecodeError:
        print(colored(f"ERROR: Invalid JSON format in '{config_path}'.", "red"))
        return {}
    except Exception as e:
        print(colored(f"ERROR: {e}", "red"))
        return {}


def get_files(exts: Sequence[str]) -> list:
    '''
    Get all files in the current directory with specific extensions. 
    Cache Files start with "~$" will not be fetched. 
    
        Parameters:
            exts (Sequence[str]): A Sequence of extensions.

        Returns:
            files (list): A list of found files path.
    '''
    current_dir = os.getcwd()
    def get_fullpath_with_ext(ext: str):
        return glob.glob(os.path.join(current_dir, ext))
    def filter_noncache_file(file: str):
        return not os.path.basename(file).startswith('~$')
    return [*filter(filter_noncache_file, chain(*map(get_fullpath_with_ext, exts)))]


def generate_filename(ext: str="xlsx") -> str:
    '''
    Generate a xlsx filename depends on current in current timezone.
    
        Parameters:
            ext (str): File Extension wanted.

        Returns:
            filename (str): Filename with give extension.
    '''
    return f"{datetime.now().replace(microsecond=0).isoformat().replace(':','_')}.{ext}"


def safe_strip(s: str):
    '''
    Do strip without exception
    
        Parameters:
            s (str): A string to do strip

        Returns:
            s (str): A string after strip
    '''
    try:
        return s.strip() if isinstance(s, str) else str(s) if s else ''
    except Exception as e:
        print(colored(f"Unknown ERROR: {e}", "red"))
        return s


def convert_excels(configs: dict, inputs: Sequence[str], output: str=None):
    '''
    Convert all given excel files by search target column upon configs.
    Result will be write to outputs excel file.
    
        Parameters:
            configs (dict): A dict of configs. Value will be considered as Field Name to search.
            inputs (Sequence[str]): A Sequence of excel files as input targets.
            outputs (str): An excel file as output target

        Returns:
            None
    '''
    # Validate configs, inputs, output
    error_msg = []
    if not all(map(lambda x: x in configs, REQUIRED_CONFIGS)):
        error_msg.append("ERROR: configs is not correct. SRC, DST, SRV are required.")
    if not inputs:
        error_msg.append("ERROR: No Excel files found in the given path.")
    if not output:
        error_msg.append("ERROR: output file is Empty")
    if error_msg:
        print(colored("\n".join(error_msg), "red"))
        return

    # Process configs
    reverse_configs = {val:key for key, val in configs.items() if key not in COMPLEX_CONFIGS}
    required_configs = {key:val for key, val in configs.items() if key in REQUIRED_CONFIGS}

    # Prepare output excel file
    output_workbook = openpyxl.Workbook()
    output_workbook.remove(output_workbook.active)

    # Process each Excel file
    for file_path in inputs:
        try:
            print(f"\n--- Processing file: {os.path.basename(file_path)} ---")
            input_workbook = openpyxl.load_workbook(file_path, read_only=True)

            # Iterate through all sheets
            for input_sheet_name in input_workbook.sheetnames:
                # Validate the sheet and Position the column
                input_sheet = input_workbook[input_sheet_name]
                rows = input_sheet.iter_rows(min_row=1, max_row=2, values_only=True)
                rows = (enumerate(map(safe_strip, row)) for row in rows)
                rows = chain.from_iterable(rows)
                cfgs = {reverse_configs[cell]:idx for idx, cell in rows if cell in reverse_configs}
                if len(cfgs.keys() & REQUIRED_CONFIGS) != len(REQUIRED_CONFIGS):
                    miss = REQUIRED_CONFIGS - cfgs.keys()
                    miss_configs = [required_configs[key] for key in miss]
                    print(f"> Sheet: [{input_sheet_name}] Skipped. " + colored(f"(Missing keywords: {miss_configs})", "yellow"))
                    continue
                # Prepare output sheet after validate
                print(colored(f"> Sheet: [{input_sheet_name}] meet at cols idx {cfgs}", "green"))
                output_sheet = output_workbook.create_sheet()
                output_sheet.column_dimensions['A'].width = 15
                output_sheet.column_dimensions['B'].width = 15
                output_sheet.column_dimensions['C'].width = 15
                output_sheet.column_dimensions['D'].width = 8
                output_sheet.column_dimensions['E'].width = 130

                # Iterate all cells
                output_row_idx = 0
                for input_row_idx, row in enumerate(input_sheet.iter_rows(min_row=2, values_only=True), 2):
                    # Validate each row
                    error_msg, tmp, values = [], [], defaultdict(str)
                    if (val:=safe_strip(row[cfgs["SRC"]])):
                        values["SRC"] = "; ".join(re.split('\n|,', val))
                    else: error_msg.append(f"Miss '{configs["SRC"]}'")
                    if (val:=safe_strip(row[cfgs["DST"]])):
                        values["DST"] = "; ".join(re.split('\n|,', val))
                    else: error_msg.append(f"Miss '{configs["DST"]}'")
                    if (val:=safe_strip(row[cfgs["SRV"]])):
                        srvs = re.split('\n|,', val)
                        srvs = [(f"TCP {srv}" if srv.isnumeric() else srv) for srv in srvs if srv]
                        values["SRV"] = "; ".join(srvs)
                        for pattern, replacement in configs["service_replace"].items():
                            values["SRV"] = values["SRV"].replace(pattern, replacement)
                    else: error_msg.append(f"Miss '{configs["SRV"]}'")
                    add_true, rm_true = safe_strip(row[cfgs["ADD"]]), safe_strip(row[cfgs["RM"]])
                    if add_true and rm_true:
                        error_msg.append(f"Have Both ('{configs["ADD"]}'&'{configs["RM"]}')")
                    elif add_true or rm_true:
                        values["ACT"] = "accept" if add_true else "remove"
                    else:
                        values["ACT"] = "" # ACT accept empty
                        # error_msg.append(f"Miss ('{configs["ADD"]}'|'{configs["RM"]}')")
                    if (val:=safe_strip(row[cfgs["USG"]])):
                        tmp.append(val)
                    if (val:=safe_strip(row[cfgs["CMT"]])):
                        tmp.append(val)
                    values["CMT"] = "|".join(tmp)
                    if error_msg and len(error_msg) == 3: # Skip empty line
                        continue
                    elif error_msg:
                        print(f"  ├─ Row {input_row_idx} Skipped. " + colored(f"{', '.join(error_msg)}", "yellow"))
                        continue

                    # Write to output_sheet
                    output_row_idx += 1
                    output_sheet.cell(row=output_row_idx, column=1, value=values["SRC"])
                    output_sheet.cell(row=output_row_idx, column=2, value=values["DST"])
                    output_sheet.cell(row=output_row_idx, column=3, value=values["SRV"])
                    output_sheet.cell(row=output_row_idx, column=4, value=values["ACT"])
                    output_sheet.cell(row=output_row_idx, column=5, value=values["CMT"])
        except openpyxl.utils.exceptions.InvalidFileException as e:
            print(colored(f"ERROR: {e}", "red"))
        # except Exception as e:
        #     print(colored(f"Unknown ERROR: {e}", "red"))
        finally:
            input_workbook.close()
    output_workbook.save(output)


# Run the script
if __name__ == "__main__":
    print(f"Current Directory: {os.getcwd()}")
    start_logging(LOG_PATH)
    keywords = get_configs(CONFIG_PATH)
    excels = get_files(EXTS)
    result = generate_filename()
    convert_excels(keywords, excels, result)
    input(colored(f"回車關閉此終端。此終端訊息皆存於 {LOG_PATH}\n", "blue"))
